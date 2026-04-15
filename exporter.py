from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def to_excel(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Организаторы"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    columns = list(df.columns)

    # Write header
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Write data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-fit column widths
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        header_len = len(str(col_name))
        max_data_len = 0
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_data_len = max(max_data_len, len(str(val)))
        width = min(max(header_len, max_data_len) + 4, 60)
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 30

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_csv(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
